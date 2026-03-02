from pyModbusTCP.client import ModbusClient
import time
import openpyxl
from openpyxl import Workbook
import os
import datetime

# --- KONFIGURASI FILE EXCEL ---
EXCEL_FILENAME = "data_sensor_tanah.xlsx"
SHEET_NAME = "Data Pembacaan Sensor"

# --- KONFIGURASI MODBUS ---
SERVER_HOST = "10.10.100.254"  # Ganti dengan IP Address HF2211 Anda
SERVER_PORT = 8899             # Ganti dengan Port TCP HF2211 Anda
SLAVE_IDS = [1, 2, 3]          # Daftar Slave ID yang akan dibaca

# --- KONFIGURASI PEMBACAAN DATA KONTINU ---
REGISTER_START_ADDRESS = 3     
REGISTER_COUNT = 1             # Total 8 register (dari alamat 0 hingga 7)
POLLING_INTERVAL = 1           # Interval total pembacaan dalam detik

# --- PETA NAMA REGISTER ---
REGISTER_MAP = {
    "kelembapan_tanah": 0,
    "suhu": 1,
    "ph_tanah": 3,  # Indeks 3
    "nitrogen": 4,
    "fosfor": 5,
    "kalium": 6,
    "salinity": 7
}

# --- PARAMETER YANG MEMBUTUHKAN SCALING 0.1 ---
SCALED_PARAMETERS = ["suhu", "kelembapan_tanah", "ph_tanah"] 

# --- FUNGSI UNTUK INISIALISASI EXCEL ---
def setup_excel():
    """Mengatur file Excel dan mengembalikan objek workbook dan worksheet."""
    if os.path.exists(EXCEL_FILENAME):
        print(f"File {EXCEL_FILENAME} ditemukan. Menambahkan data ke worksheet '{SHEET_NAME}'.")
        wb = openpyxl.load_workbook(EXCEL_FILENAME)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            headers = ["Timestamp", "Slave ID"] + [name.replace('_', ' ').title() for name in REGISTER_MAP.keys()]
            ws.append(headers) # Tulis header jika sheet baru dibuat
        else:
            ws = wb[SHEET_NAME]
    else:
        print(f"File {EXCEL_FILENAME} tidak ditemukan. Membuat file baru.")
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        # Tulis header kolom
        headers = ["Timestamp", "Slave ID"] + [name.replace('_', ' ').title() for name in REGISTER_MAP.keys()]
        ws.append(headers)
        
    return wb, ws

# --- INISIALISASI DAN LOOP UTAMA ---

# Inisialisasi Klien Modbus TCP
client = ModbusClient(
    host=SERVER_HOST, 
    port=SERVER_PORT, 
    auto_open=True, 
    auto_close=False
)

# Inisialisasi Excel
wb, ws = setup_excel()

# ... (kode sebelum try) ...

print(f"Mencoba koneksi ke Modbus Gateway {SERVER_HOST}:{SERVER_PORT}...")

try: # Baris ini adalah awal dari blok try (sekitar baris 74 di kode asli Anda)
    if client.open():
        print("Koneksi Modbus berhasil dibuka. Memulai pembacaan kontinu...")
        # ... (Siklus Pembacaan Kontinu) ...
    else:
        print("Gagal membuka koneksi Modbus. Periksa gateway.")
        
    while True:
        # ... (Loop pembacaan) ...
        time.sleep(POLLING_INTERVAL)

except KeyboardInterrupt: # Menangkap error pengguna menekan Ctrl+C
    print("\nProgram dihentikan oleh pengguna (Ctrl+C).")
except Exception as e: # Menangkap error umum lainnya
    print(f"\nTerjadi kesalahan tak terduga: {e}")
    
finally: # Kode ini dijalankan terlepas dari keberhasilan try atau except
    if 'wb' in locals() or 'wb' in globals():
        print(f"Menyimpan file Excel terakhir...")
        wb.save(EXCEL_FILENAME)
        
    if client.is_open:
        client.close()
        print("Koneksi Modbus ditutup.")