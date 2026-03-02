from pyModbusTCP.client import ModbusClient
import time
import openpyxl
from openpyxl import Workbook
import os
import datetime
from typing import Dict, Any, Tuple

# --- KONFIGURASI FILE EXCEL ---
EXCEL_FILENAME = "data_monitoring_tanah.xlsx"
SHEET_NAME = "Data Analisis Sensor"

# --- KONFIGURASI MODBUS ---
SERVER_HOST = "10.10.100.254"
SERVER_PORT = 8899
SLAVE_IDS = [1, 2, 3]

# --- KONFIGURASI PEMBACAAN DATA KONTINU ---
REGISTER_START_ADDRESS = 0     
REGISTER_COUNT = 8             
POLLING_INTERVAL = 1           

# --- PETA NAMA REGISTER & SKALA ---
REGISTER_MAP = {
    "kelembapan_tanah": 0,
    "suhu": 1,
    "ph_tanah": 3,
    "nitrogen": 4,
    "fosfor": 5,
    "kalium": 6,
    "salinity": 7
}

SCALED_PARAMETERS = ["suhu", "kelembapan_tanah", "ph_tanah"] 

# --- AMBANG BATAS KESEHATAN TANAH (Dapat Disesuaikan) ---
HEALTH_LIMITS: Dict[str, Tuple[float, float]] = {
    "kelembapan_tanah": (30.0, 60.0), 
    "suhu": (20.0, 35.0),             
    "ph_tanah": (5.5, 7.0),           
    "nitrogen": (50, 100),            
    "fosfor": (15, 40),               
    "kalium": (100, 200),             
    "salinity": (0.0, 2.0)            
}

# --- FUNGSI EXCEL (Diambil dari Modifikasi Sebelumnya) ---
def setup_excel() -> Tuple[Workbook, Any]:
    """Mengatur file Excel dan mengembalikan objek workbook dan worksheet."""
    if os.path.exists(EXCEL_FILENAME):
        wb = openpyxl.load_workbook(EXCEL_FILENAME)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
        else:
            ws = wb[SHEET_NAME]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
    
    # Tulis header hanya jika worksheet kosong
    if ws.max_row == 1 and ws['A1'].value is None:
        headers = ["Timestamp", "Slave ID"] + [name.replace('_', ' ').title() for name in REGISTER_MAP.keys()] + ["Status Kesehatan", "Detail Masalah"]
        ws.append(headers)
        
    return wb, ws

# --- FUNGSI ANALISIS KESEHATAN (Diperbaiki untuk konsol) ---
def get_health_status(data: Dict[str, Any]) -> Tuple[str, str, str]:
    """Menentukan status kesehatan berdasarkan ambang batas."""
    
    overall_status = "SEHAT"
    issues = []
    emoji = "✅" # Default Sehat
    
    for name, value in data.items():
        if name in HEALTH_LIMITS:
            min_ideal, max_ideal = HEALTH_LIMITS[name]
            
            if value < min_ideal:
                issues.append(f"{name.replace('_', ' ').title()} ({value}) terlalu RENDAH (Min: {min_ideal})")
            elif value > max_ideal:
                issues.append(f"{name.replace('_', ' ').title()} ({value}) terlalu TINGGI (Max: {max_ideal})")
        
    if len(issues) >= 3:
        overall_status = "KRITIS"
        emoji = "❌"
    elif len(issues) > 0:
        overall_status = "PERLU PERHATIAN"
        emoji = "⚠️"
    
    issues_str = "\n  - ".join(issues)
    return overall_status, issues_str, emoji

# --- INISIALISASI UTAMA ---

client = ModbusClient(
    host=SERVER_HOST, 
    port=SERVER_PORT, 
    auto_open=True, 
    auto_close=False
)

# Inisialisasi Excel
wb, ws = setup_excel()

# ... (Kode di atas baris 109 - Inisialisasi Klien Modbus & Excel) ...

print(f"Mencoba koneksi ke Modbus Gateway {SERVER_HOST}:{SERVER_PORT}...")

try: # BARIS 109: Tidak ada indentasi
    if client.open():
        print("Koneksi Modbus berhasil dibuka. Memulai monitoring kesehatan tanah...")
    else:
        print("Gagal membuka koneksi Modbus. Periksa gateway.")
        exit()
        
    # --- Loop Pembacaan Kontinu ---
    while True:
        timestamp_cycle = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"\n--- {timestamp_cycle} - MONITORING KESEHATAN TANAH ---")
        
        for unit_id in SLAVE_IDS:
            client.unit_id = unit_id
            registers = client.read_holding_registers(REGISTER_START_ADDRESS, REGISTER_COUNT)

            if registers:
                # 1. PETAKAN DAN SKALAKAN DATA
                mapped_data: Dict[str, float | int] = {}
                excel_row_data = [timestamp_cycle, unit_id]
                
                for name, index in REGISTER_MAP.items():
                    raw_value = registers[index]
                    scaled_value: float | int = raw_value
                    
                    if name in SCALED_PARAMETERS:
                        scaled_value = raw_value / 10.0 
                    
                    mapped_data[name] = scaled_value
                    excel_row_data.append(scaled_value) # Tambahkan nilai ke baris Excel

                # 2. ANALISIS KESEHATAN
                status, issues_detail, emoji = get_health_status(mapped_data)
                
                # Tambahkan status dan detail ke baris Excel
                excel_row_data.append(status)
                excel_row_data.append(issues_detail.replace("\n  - ", "; ")) 
                
                # Tulis ke Excel
                ws.append(excel_row_data)

                # 3. CETAK HASIL KE KONSOL
                print(f"[{emoji} {status}] Data Slave ID {unit_id}:")
                
                # Cetak semua nilai parameter
                for name, value in mapped_data.items():
                    display_name = name.replace('_', ' ').title()
                    format_str = f":.1f" if name in SCALED_PARAMETERS else f""
                    print(f"  > {display_name}: {{value{format_str}}}".format(value=value))

                # Cetak detail masalah
                if issues_detail:
                    print(f"  [MASALAH TERDETEKSI]: {issues_detail}")
                
            else:
                print(f"[ID {unit_id}] GAGAL membaca register (Timeout/Koneksi Serial).")
                if not client.is_open:
                    client.open()

        # Simpan file Excel sekali setelah semua unit dibaca
        wb.save(EXCEL_FILENAME)
        print(f"Data siklus berhasil disimpan ke {EXCEL_FILENAME}")

        time.sleep(POLLING_INTERVAL)

except KeyboardInterrupt: # HARUS SEJAJAR DENGAN try
    print("\nProgram dihentikan oleh pengguna (Ctrl+C).")
except Exception as e: # HARUS SEJAJAR DENGAN try
    print(f"\nTerjadi kesalahan tak terduga: {e}")
    
finally: # HARUS SEJAJAR DENGAN try
    # Pastikan file Excel disimpan dan koneksi ditutup
    if 'wb' in locals() or 'wb' in globals():
        try:
            wb.save(EXCEL_FILENAME)
            print(f"File {EXCEL_FILENAME} berhasil disimpan terakhir kali.")
        except Exception as save_e:
            print(f"⚠️ GAGAL MENYIMPAN FILE EXCEL FINAL: {save_e}")

    if client.is_open:
        client.close()
        print("Koneksi Modbus ditutup.")